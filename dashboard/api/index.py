"""
Metabase proxy API — deployed as a Vercel Python serverless function.
Routes: /api/login, /api/databases, /api/query, /api/export/excel, /api/export/csv
"""

import csv
import io
import json
import logging

import httpx
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(title="Metabase Proxy API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

METABASE_INSTANCES = {
    "regular": "https://metabase.auditoria.ai",
    "enterprise": "https://metabase-ent1.auditoria.ai",
}

# ---------------------------------------------------------------------------
# Request models
# ---------------------------------------------------------------------------

class LoginRequest(BaseModel):
    instance: str
    username: str
    password: str


class QueryRequest(BaseModel):
    instance: str
    session_token: str
    database_id: int
    sql: str


class ExportRequest(BaseModel):
    columns: list[str]
    rows: list[list]
    format: str = "excel"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def get_base_url(instance: str) -> str:
    url = METABASE_INSTANCES.get(instance)
    if not url:
        raise HTTPException(status_code=400, detail=f"Unknown instance '{instance}'. Use 'regular' or 'enterprise'.")
    return url


def mb_headers(session_token: str) -> dict:
    return {
        "Content-Type": "application/json",
        "X-Metabase-Session": session_token,
    }


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------

@app.post("/api/login")
async def login(req: LoginRequest):
    base_url = get_base_url(req.instance)
    async with httpx.AsyncClient(verify=False, timeout=30) as client:
        try:
            resp = await client.post(
                f"{base_url}/api/session",
                json={"username": req.username, "password": req.password},
                headers={"Content-Type": "application/json"},
            )
        except httpx.ConnectError as e:
            raise HTTPException(status_code=503, detail=f"Cannot reach {base_url}: {e}")
        except httpx.TimeoutException:
            raise HTTPException(status_code=504, detail="Connection timed out")

    if resp.status_code == 401:
        raise HTTPException(status_code=401, detail="Invalid username or password")
    if resp.status_code != 200:
        raise HTTPException(status_code=resp.status_code, detail=f"Metabase error: {resp.text}")

    data = resp.json()
    token = data.get("id") or data.get("token")
    if not token:
        raise HTTPException(status_code=500, detail="No session token in Metabase response")

    return {"session_token": token, "instance": req.instance, "base_url": base_url}


# ---------------------------------------------------------------------------
# Databases
# ---------------------------------------------------------------------------

@app.get("/api/databases")
async def list_databases(instance: str, session_token: str):
    base_url = get_base_url(instance)
    async with httpx.AsyncClient(verify=False, timeout=30) as client:
        try:
            resp = await client.get(
                f"{base_url}/api/database",
                headers=mb_headers(session_token),
                params={"include_tables": "true"},
            )
        except httpx.ConnectError as e:
            raise HTTPException(status_code=503, detail=str(e))
        except httpx.TimeoutException:
            raise HTTPException(status_code=504, detail="Timed out fetching databases")

    if resp.status_code == 401:
        raise HTTPException(status_code=401, detail="Session expired or invalid")
    if resp.status_code != 200:
        raise HTTPException(status_code=resp.status_code, detail=resp.text)

    data = resp.json()
    databases = data if isinstance(data, list) else data.get("data", data)

    result = []
    for db in databases:
        tables = db.get("tables", [])
        result.append({
            "id": db["id"],
            "name": db["name"],
            "engine": db.get("engine", ""),
            "tables": [
                {"id": t["id"], "name": t["name"], "schema": t.get("schema", "")}
                for t in tables
            ],
        })
    return result


# ---------------------------------------------------------------------------
# Query — returns all rows as JSON (no SSE, compatible with Vercel)
# ---------------------------------------------------------------------------

@app.post("/api/query")
async def run_query(req: QueryRequest):
    base_url = get_base_url(req.instance)
    payload = {
        "database": req.database_id,
        "type": "native",
        "native": {"query": req.sql},
    }

    async with httpx.AsyncClient(verify=False, timeout=300) as client:
        try:
            resp = await client.post(
                f"{base_url}/api/dataset",
                json=payload,
                headers=mb_headers(req.session_token),
            )
        except httpx.ConnectError as e:
            raise HTTPException(status_code=503, detail=f"Cannot reach {base_url}: {e}")
        except httpx.TimeoutException:
            raise HTTPException(status_code=504, detail="Query timed out (300s limit)")

    if resp.status_code == 401:
        raise HTTPException(status_code=401, detail="Session expired. Please log in again.")
    if resp.status_code == 403:
        raise HTTPException(status_code=403, detail="Permission denied for this database.")
    if resp.status_code not in (200, 202):
        try:
            err = resp.json().get("message", resp.text)
        except Exception:
            err = resp.text
        raise HTTPException(status_code=resp.status_code, detail=f"Metabase error: {err}")

    try:
        data = resp.json()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to parse Metabase response: {e}")

    if data.get("error"):
        raise HTTPException(status_code=400, detail=str(data["error"]))

    result_data = data.get("data", data)
    cols = result_data.get("cols") or result_data.get("results_metadata", {}).get("columns", [])
    rows = result_data.get("rows") or []
    columns = [c.get("display_name") or c.get("name", f"col_{i}") for i, c in enumerate(cols)]

    return {"columns": columns, "rows": rows, "total": len(rows)}


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

@app.post("/api/export/excel")
async def export_excel(req: ExportRequest):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = Workbook()
    ws = wb.active
    ws.title = "Query Results"

    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col_idx, col_name in enumerate(req.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(req.rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_idx in range(1, len(req.columns) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(req.columns[col_idx - 1]))
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="metabase_results.xlsx"'},
    )


@app.post("/api/export/csv")
async def export_csv(req: ExportRequest):
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(req.columns)
    writer.writerows(req.rows)
    buf.seek(0)

    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="metabase_results.csv"'},
    )


# ---------------------------------------------------------------------------
# Health check
# ---------------------------------------------------------------------------

@app.get("/api")
async def health():
    return JSONResponse({"status": "Metabase API running"})
